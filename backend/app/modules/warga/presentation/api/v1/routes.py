# warga/presentation/api/v1/routes.py
import uuid as _uuid
from datetime import datetime, timezone
from uuid import UUID

from app.core.database import get_db
from app.core.dependencies import get_current_user, require_admin
from app.core.exceptions import EntityNotFoundError
from app.modules.warga.application.schemas import (AddAnggotaRequest,
                                                   AdminCreateResidentRequest,
                                                   AdminCreateResidentResponse,
                                                   AdminUpdateResidentRequest,
                                                   ChangeLogEntry,
                                                   ChangeRequestItem,
                                                   RegisterResidentRequest,
                                                   ReviewChangeRequestBody,
                                                   SubmitChangeRequestBody,
                                                   SubmitChangeRequestResponse)
from app.modules.warga.application.use_cases.import_residents import (
    bulk_create, parse_excel)
from app.modules.warga.application.use_cases.register_resident import \
    RegisterResident
from app.modules.warga.application.use_cases.verify_resident import \
    VerifyResident
from app.modules.warga.infrastructure.models import (
    ResidentChangeLogModel, ResidentChangeRequestModel, ResidentModel)
from app.modules.warga.infrastructure.repository import PgResidentRepository
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from fastapi.responses import StreamingResponse
from sqlalchemy import desc, select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession

router = APIRouter()


@router.post("/warga", status_code=201, tags=["Warga"])
async def register_resident(
    body: RegisterResidentRequest,
    current_user: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    resident = await RegisterResident(PgResidentRepository(db)).execute(
        rt_group_id=body.rt_group_id,
        user_id=UUID(current_user["user_id"]),
        full_name=body.full_name, phone=body.phone,
        street=body.street, rt_number=body.rt_number,
        rw_number=body.rw_number, kelurahan=body.kelurahan,
        kecamatan=body.kecamatan, kota=body.kota,
        block=body.block, unit_number=body.unit_number,
        ownership_type=body.ownership_type, member_count=body.member_count,
    )
    return {"id": str(resident.id), "status": resident.status}


# ═══════════════════════════════════════════════════════════════════════════════
# === ADDED — POST /warga/admin-create
# Tambah Warga: Ketua RT manually adds a warga's data — no login account
# required. Creates a "ghost" resident record (user_id=None, status=ACTIVE).
# ═══════════════════════════════════════════════════════════════════════════════

@router.post("/warga/admin-create", status_code=201,
              response_model=AdminCreateResidentResponse, tags=["Warga"])
async def admin_create_resident(
    body:         AdminCreateResidentRequest,
    current_user: dict = Depends(require_admin),
    db:           AsyncSession = Depends(get_db),
):
    """
    Ketua RT manually adds a warga's data — no login account required.

    Only full_name + phone are mandatory. NIK/No KK/status_keluarga/
    alamat_ktp/alamat_domisili are optional and can be filled in later
    via PATCH /warga/{resident_id}/admin-update.

    The resulting record has user_id=None ("ghost" resident). Linking
    this record to a real account when the warga self-registers is a
    future enhancement — not handled in v1.
    """
    from app.modules.iam.infrastructure.models import RTGroupModel
    from app.modules.warga.domain.entities import Resident, StatusKeluarga

    admin_id = _uuid.UUID(current_user["user_id"])

    rt_result = await db.execute(
        select(RTGroupModel).where(RTGroupModel.admin_user_id == admin_id)
    )
    rt_group = rt_result.scalar_one_or_none()
    if not rt_group:
        raise HTTPException(status_code=404, detail="RT group tidak ditemukan untuk akun ini")

    repo = PgResidentRepository(db)
    resident = Resident.create_by_admin(
        rt_group_id=rt_group.id,
        full_name=body.full_name,
        phone=body.phone,
        added_by_user_id=admin_id,
        nik=body.nik,
        no_kk=body.no_kk,
        status_keluarga=StatusKeluarga(body.status_keluarga) if body.status_keluarga else None,
        alamat_ktp=body.alamat_ktp,
        alamat_domisili=body.alamat_domisili,
    )

    try:
        await repo.save(resident)
        await db.commit()
    except IntegrityError:
        await db.rollback()
        raise HTTPException(
            status_code=409,
            detail="Gagal menyimpan — NIK atau No. KK mungkin sudah terdaftar"
        )

    return AdminCreateResidentResponse(
        id=str(resident.id),
        full_name=resident.full_name,
        phone=resident.phone,
        status=resident.status.value,
        message=f"{resident.full_name} berhasil ditambahkan",
    )


@router.get("/warga/rt/{rt_group_id}", tags=["Warga"])
async def list_residents(
    rt_group_id: UUID,
    status_filter: str = None,
    current_user: dict = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    repo = PgResidentRepository(db)
    from app.modules.warga.domain.entities import ResidentStatus
    status_enum = ResidentStatus(status_filter) if status_filter else None
    residents = await repo.get_by_rt_group(rt_group_id, status=status_enum)
    return [{"id": str(r.id), "full_name": r.full_name, "phone": r.phone,
             "block_unit": r.block_unit_display, "status": r.status,
             "member_count": r.member_count} for r in residents]



@router.get("/warga/my-profile", tags=["Warga"])
async def get_my_profile(
    current_user: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Returns the resident profile for the logged-in warga."""
    repo     = PgResidentRepository(db)
    resident = await repo.get_by_user_id(UUID(current_user["user_id"]))
    if not resident:
        return {}
    return {
        "nik":             resident.nik,
        "no_kk":           resident.no_kk,
        "tanggal_lahir":   resident.tanggal_lahir.isoformat() if resident.tanggal_lahir else None,
        "tempat_lahir":    resident.tempat_lahir,
        "jenis_kelamin":   resident.jenis_kelamin.value if resident.jenis_kelamin else None,
        "agama":           resident.agama.value          if resident.agama         else None,
        "pekerjaan":       resident.pekerjaan.value      if resident.pekerjaan     else None,
        "status_kawin":    resident.status_kawin.value   if resident.status_kawin  else None,
        "status_tinggal":  resident.status_tinggal.value if resident.status_tinggal else None,
        "status_keluarga": resident.status_keluarga.value if resident.status_keluarga else None,
        "kepala_keluarga":     resident.kepala_keluarga,
        "alamat_ktp":          resident.alamat_ktp,
        "pendidikan_terakhir": resident.pendidikan_terakhir.value if resident.pendidikan_terakhir else None,
        "kewarganegaraan":     resident.kewarganegaraan.value if resident.kewarganegaraan else None,
        "hubungan_dengan_kk":  resident.hubungan_dengan_kk.value if resident.hubungan_dengan_kk else None,
    }

@router.get("/warga/user/{user_id}/profile", tags=["Warga"])
async def get_warga_full_profile(
    user_id:      UUID,
    current_user: dict = Depends(require_admin),
    db:           AsyncSession = Depends(get_db),
):
    """
    Returns full resident profile for a given user_id.
    Used by clicking a warga row in Data Warga page.
    """
    repo     = PgResidentRepository(db)
    resident = await repo.get_by_user_id(user_id)
    if not resident:
        raise HTTPException(status_code=404, detail="Data warga tidak ditemukan")

    result = _resident_detail_from_entity(resident)

    # If they have a no_kk, fetch all KK members too
    if resident.no_kk:
        from sqlalchemy import select as sa_select
        kk_result = await db.execute(
            sa_select(ResidentModel)
            .where(
                ResidentModel.no_kk == resident.no_kk,
                ResidentModel.id != resident.id,
            )
            .order_by(ResidentModel.kepala_keluarga.desc())
        )
        other_members = kk_result.scalars().all()
        result["kk_members"] = [_resident_detail(r) for r in other_members]
    else:
        result["kk_members"] = []

    return result


def _resident_detail(row) -> dict:
    """Convert ResidentModel row to detail dict."""
    return {
        "id":                  str(row.id),
        "full_name":           row.full_name,
        "nik":                 row.nik,
        "no_kk":               row.no_kk,
        "tanggal_lahir":       row.tanggal_lahir.isoformat() if row.tanggal_lahir else None,
        "tempat_lahir":        row.tempat_lahir,
        "jenis_kelamin":       row.jenis_kelamin,
        "agama":               row.agama,
        "pekerjaan":           row.pekerjaan,
        "status_kawin":        row.status_kawin,
        "status_tinggal":      row.status_tinggal,
        "status_keluarga":     row.status_keluarga,
        "hubungan_dengan_kk":  row.hubungan_dengan_kk,
        "kepala_keluarga":     row.kepala_keluarga,
        "pendidikan_terakhir": row.pendidikan_terakhir,
        "kewarganegaraan":     row.kewarganegaraan,
        "alamat_ktp":          row.alamat_ktp,
        "phone":               row.phone,
        "block_unit":          f"Blok {row.block} No. {row.unit_number}" if row.block else None,
    }


def _resident_detail_from_entity(r) -> dict:
    """Convert Resident entity to detail dict."""
    return {
        "id":                  str(r.id),
        "full_name":           r.full_name,
        "nik":                 r.nik,
        "no_kk":               r.no_kk,
        "tanggal_lahir":       r.tanggal_lahir.isoformat() if r.tanggal_lahir else None,
        "tempat_lahir":        r.tempat_lahir,
        "jenis_kelamin":       r.jenis_kelamin.value if r.jenis_kelamin else None,
        "agama":               r.agama.value if r.agama else None,
        "pekerjaan":           r.pekerjaan.value if r.pekerjaan else None,
        "status_kawin":        r.status_kawin.value if r.status_kawin else None,
        "status_tinggal":      r.status_tinggal.value if r.status_tinggal else None,
        "status_keluarga":     r.status_keluarga.value if r.status_keluarga else None,
        "hubungan_dengan_kk":  r.hubungan_dengan_kk.value if r.hubungan_dengan_kk else None,
        "kepala_keluarga":     r.kepala_keluarga,
        "pendidikan_terakhir": r.pendidikan_terakhir.value if r.pendidikan_terakhir else None,
        "kewarganegaraan":     r.kewarganegaraan.value if r.kewarganegaraan else "WNI",
        "alamat_ktp":          r.alamat_ktp,
        "phone":               r.phone,
        "block_unit":          r.block_unit_display,
    }

@router.get("/warga/kk/{no_kk}", tags=["Warga"])
async def get_kk_members(
    no_kk:        str,
    current_user: dict = Depends(require_admin),
    db:           AsyncSession = Depends(get_db),
):
    """
    Returns all residents sharing the same no_kk (Kartu Keluarga).
    Used by the KK modal in Data Warga page.
    """
    from app.modules.iam.infrastructure.models import UserModel
    from sqlalchemy import select as sa_select

    result = await db.execute(
        sa_select(ResidentModel)
        .where(ResidentModel.no_kk == no_kk)
        .order_by(ResidentModel.kepala_keluarga.desc())
    )
    residents = result.scalars().all()

    if not residents:
        return []

    repo = PgResidentRepository(db)
    return [_resident_detail(r) for r in residents]

@router.get("/warga/my-keluarga", tags=["Warga"])
async def get_my_keluarga(
    current_user: dict = Depends(get_current_user),
    db:           AsyncSession = Depends(get_db),
):
    """Returns all anggota KK added by the logged-in warga."""
    from sqlalchemy import select as sa_select
    repo    = PgResidentRepository(db)
    user_id = UUID(current_user["user_id"])
    me      = await repo.get_by_user_id(user_id)

    if not me:
        return {"kepala": None, "anggota": [], "no_kk": None}

    result = await db.execute(
        sa_select(ResidentModel)
        .where(
            ResidentModel.added_by_user_id == user_id,
            ResidentModel.is_anggota_kk   == True,
        )
        .order_by(ResidentModel.created_at)
    )
    anggota_rows = result.scalars().all()

    return {
        "no_kk":   me.no_kk,
        "kepala":  _resident_detail_from_entity(me),
        "anggota": [_resident_detail(r) for r in anggota_rows],
    }


@router.post("/warga/anggota", status_code=201, tags=["Warga"])
async def add_anggota_kk(
    body:         AddAnggotaRequest,
    current_user: dict = Depends(get_current_user),
    db:           AsyncSession = Depends(get_db),
):
    """Warga (Kepala KK) adds a family member without a user account."""
    import uuid as _uuid
    from datetime import date as _date
    from datetime import datetime, timezone

    repo    = PgResidentRepository(db)
    user_id = UUID(current_user["user_id"])
    me      = await repo.get_by_user_id(user_id)

    if not me:
        raise HTTPException(status_code=403,
            detail="Anda belum terdaftar sebagai warga RT")

    if not me.no_kk:
        raise HTTPException(status_code=422,
            detail="Lengkapi No. KK di Profil Saya terlebih dahulu "
                   "sebelum menambah anggota keluarga")

    tanggal_lahir = None
    if body.tanggal_lahir:
        try:
            tanggal_lahir = _date.fromisoformat(body.tanggal_lahir)
        except ValueError:
            pass

    now = datetime.now(timezone.utc)
    anggota_row = ResidentModel(
        id               = _uuid.uuid4(),
        rt_group_id      = me.rt_group_id,
        user_id          = None,
        full_name        = body.full_name.strip(),
        phone            = body.phone or "",
        no_kk            = me.no_kk,
        nik              = body.nik,
        tanggal_lahir    = tanggal_lahir,
        tempat_lahir     = body.tempat_lahir,
        jenis_kelamin    = body.jenis_kelamin,
        agama            = body.agama,
        pekerjaan        = body.pekerjaan,
        status_kawin     = body.status_kawin,
        status_tinggal   = body.status_tinggal or "TETAP",
        hubungan_dengan_kk  = body.hubungan_dengan_kk,
        pendidikan_terakhir = body.pendidikan_terakhir,
        kewarganegaraan  = body.kewarganegaraan or "WNI",
        kepala_keluarga  = False,
        is_anggota_kk    = True,
        added_by_user_id = user_id,
        status           = "active",
        ownership_type   = me.ownership_type.value if me.ownership_type else "owner",
        street           = me.street or "",
        rt_number        = me.rt_number or "",
        rw_number        = me.rw_number or "",
        kelurahan        = me.kelurahan or "",
        kecamatan        = me.kecamatan or "",
        kota             = me.kota or "",
        block            = me.block or "",
        unit_number      = me.unit_number or "",
        member_count     = 1,
        created_at       = now,
        updated_at       = now,
    )
    db.add(anggota_row)
    await db.flush()
    await db.commit()
    await db.refresh(anggota_row)

    return {
        "id":        str(anggota_row.id),
        "full_name": anggota_row.full_name,
        "no_kk":     anggota_row.no_kk,
        "hubungan":  anggota_row.hubungan_dengan_kk,
        "message":   f"{body.full_name} berhasil ditambahkan sebagai anggota KK",
    }


@router.delete("/warga/anggota/{anggota_id}", tags=["Warga"])
async def delete_anggota_kk(
    anggota_id:   UUID,
    current_user: dict = Depends(get_current_user),
    db:           AsyncSession = Depends(get_db),
):
    """Kepala KK removes a family member they added."""
    from sqlalchemy import select as sa_select
    user_id = UUID(current_user["user_id"])

    result = await db.execute(
        sa_select(ResidentModel)
        .where(
            ResidentModel.id               == anggota_id,
            ResidentModel.added_by_user_id == user_id,
            ResidentModel.is_anggota_kk    == True,
        )
    )
    row = result.scalar_one_or_none()
    if not row:
        raise HTTPException(status_code=404,
            detail="Anggota tidak ditemukan atau bukan milik Anda")

    await db.delete(row)
    await db.commit()
    return {"message": f"{row.full_name} berhasil dihapus dari daftar KK"}


@router.get("/warga/export/{rt_group_id}", tags=["Warga"])
async def export_warga_data(
    rt_group_id:  UUID,
    current_user: dict = Depends(require_admin),
    db:           AsyncSession = Depends(get_db),
):
    """Returns full resident data for Excel export."""
    from sqlalchemy import select as sa_select

    result = await db.execute(
        sa_select(ResidentModel)
        .where(
            ResidentModel.rt_group_id == rt_group_id,
            ResidentModel.status      == "active",
        )
        .order_by(
            ResidentModel.no_kk.nullslast(),
            ResidentModel.kepala_keluarga.desc(),
            ResidentModel.full_name,
        )
    )
    residents = result.scalars().all()

    return [
        {
            "id":                  str(r.id),
            "full_name":           r.full_name,
            "nik":                 r.nik,
            "no_kk":               r.no_kk,
            "tanggal_lahir":       r.tanggal_lahir.isoformat() if r.tanggal_lahir else None,
            "tempat_lahir":        r.tempat_lahir,
            "jenis_kelamin":       r.jenis_kelamin,
            "agama":               r.agama,
            "pekerjaan":           r.pekerjaan,
            "status_kawin":        r.status_kawin,
            "status_tinggal":      r.status_tinggal,
            "hubungan_dengan_kk":  r.hubungan_dengan_kk,
            "kepala_keluarga":     r.kepala_keluarga,
            "pendidikan_terakhir": r.pendidikan_terakhir,
            "kewarganegaraan":     r.kewarganegaraan or "WNI",
            "alamat_ktp":          r.alamat_ktp,
            "phone":               r.phone,
            "block_unit":          f"Blok {r.block} No. {r.unit_number}" if r.block else None,
            "is_anggota_kk":       r.is_anggota_kk,
        }
        for r in residents
    ]


@router.get("/warga/statistik/{rt_group_id}", tags=["Warga"])
async def get_statistik_demografis(
    rt_group_id:  UUID,
    current_user: dict = Depends(require_admin),
    db:           AsyncSession = Depends(get_db),
):
    from datetime import date

    from sqlalchemy import select as sa_select

    result = await db.execute(
        sa_select(ResidentModel).where(
            ResidentModel.rt_group_id == rt_group_id,
            ResidentModel.status      == "active",
        )
    )
    all_residents = result.scalars().all()
    total = len(all_residents)

    if total == 0:
        return {"total_warga": 0, "total_kk": 0, "kepala_keluarga": 0,
                "jenis_kelamin": [], "agama": [], "pendidikan": [],
                "pekerjaan": [], "status_tinggal": [], "usia": [],
                "kewarganegaraan": []}

    def count_by(field):
        counts = {}
        for r in all_residents:
            val = getattr(r, field, None) or "Tidak Diisi"
            counts[val] = counts.get(val, 0) + 1
        return [{"name": k, "value": v}
                for k, v in sorted(counts.items(), key=lambda x: -x[1])]

    today = date.today()
    usia_buckets = {"Balita (0-4)": 0, "Anak (5-11)": 0,
                    "Remaja (12-18)": 0, "Dewasa (19-59)": 0,
                    "Lansia (60+)": 0, "Tidak Diisi": 0}
    for r in all_residents:
        if not r.tanggal_lahir:
            usia_buckets["Tidak Diisi"] += 1
            continue
        age = (today - r.tanggal_lahir).days // 365
        if age <= 4:    usia_buckets["Balita (0-4)"] += 1
        elif age <= 11: usia_buckets["Anak (5-11)"] += 1
        elif age <= 18: usia_buckets["Remaja (12-18)"] += 1
        elif age <= 59: usia_buckets["Dewasa (19-59)"] += 1
        else:           usia_buckets["Lansia (60+)"] += 1

    PENDIDIKAN_ORDER = ["TIDAK SEKOLAH","BELUM SEKOLAH","SD","SMP","SMA",
                        "SMK","D3","S1","S2","S3","LAINNYA","Tidak Diisi"]
    pend_counts = {}
    for r in all_residents:
        val = r.pendidikan_terakhir or "Tidak Diisi"
        pend_counts[val] = pend_counts.get(val, 0) + 1

    return {
        "total_warga":     total,
        "total_kk":        len({r.no_kk for r in all_residents if r.no_kk}),
        "kepala_keluarga": sum(1 for r in all_residents if r.kepala_keluarga),
        "jenis_kelamin":   count_by("jenis_kelamin"),
        "agama":           count_by("agama"),
        "pendidikan":      [{"name": k, "value": pend_counts[k]}
                            for k in PENDIDIKAN_ORDER if k in pend_counts],
        "pekerjaan":       count_by("pekerjaan"),
        "status_tinggal":  count_by("status_tinggal"),
        "usia":            [{"name": k, "value": v}
                            for k, v in usia_buckets.items() if v > 0],
        "kewarganegaraan": count_by("kewarganegaraan"),
    }


@router.get("/warga/{resident_id}", tags=["Warga"])
async def get_resident(
    resident_id: UUID,
    current_user: dict = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    resident = await PgResidentRepository(db).get_by_id(resident_id)
    if not resident:
        raise HTTPException(status_code=404, detail="Warga tidak ditemukan")
    return {"id": str(resident.id), "full_name": resident.full_name,
            "phone": resident.phone, "block_unit": resident.block_unit_display,
            "status": resident.status, "kk_file_url": resident.kk_file_url,
            "ktp_file_url": resident.ktp_file_url}


@router.patch("/warga/{resident_id}/verify", tags=["Warga"])
async def verify_resident(
    resident_id: UUID,
    current_user: dict = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    try:
        resident = await VerifyResident(PgResidentRepository(db)).execute(
            resident_id=resident_id, verified_by=UUID(current_user["user_id"])
        )
        return {"id": str(resident.id), "status": resident.status}
    except EntityNotFoundError as e:
        raise HTTPException(status_code=404, detail=e.message)
    

FIELD_LABELS: dict[str, str] = {
    "full_name":           "Nama Lengkap",
    "phone":               "Nomor HP",
    "nik":                 "NIK",
    "no_kk":               "Nomor KK",
    "tanggal_lahir":       "Tanggal Lahir",
    "tempat_lahir":        "Tempat Lahir",
    "jenis_kelamin":       "Jenis Kelamin",
    "agama":               "Agama",
    "pekerjaan":           "Pekerjaan",
    "status_kawin":        "Status Perkawinan",
    "status_tinggal":      "Status Tinggal",
    "status_keluarga":     "Status dalam Keluarga",
    "kepala_keluarga":     "Kepala Keluarga",
    "alamat_ktp":          "Alamat KTP",
    "pendidikan_terakhir": "Pendidikan Terakhir",
    "kewarganegaraan":     "Kewarganegaraan",
    "hubungan_dengan_kk":  "Hubungan dengan KK",
}


# ═══════════════════════════════════════════════════════════════════════════════
# PATCH /warga/{resident_id}/admin-update
# Ketua RT full-authority profile update — immediate, no approval needed
# ═══════════════════════════════════════════════════════════════════════════════

@router.patch("/warga/{resident_id}/admin-update", tags=["Warga"])
async def admin_update_resident(
    resident_id:  UUID,
    body:         AdminUpdateResidentRequest,
    current_user: dict = Depends(require_admin),
    db:           AsyncSession = Depends(get_db),
):
    """
    Ketua RT updates any warga profile field immediately.
    Every changed field is logged to resident_change_logs with
    old value, new value, who changed it, and when.

    Only fields explicitly provided in the request body are updated.
    Fields set to None in the JSON are ignored (partial update).
    Fields explicitly sent as null overwrite to null — use with care.
    """
    import uuid as _uuid
    from datetime import datetime, timezone

    from app.modules.iam.infrastructure.models import UserModel
    from app.modules.warga.domain.entities import (Agama, HubunganDenganKK,
                                                   JenisKelamin,
                                                   Kewarganegaraan, Pekerjaan,
                                                   PendidikanTerakhir,
                                                   StatusKawin, StatusKeluarga,
                                                   StatusTinggal)
    from sqlalchemy import insert as sa_insert
    from sqlalchemy import select as sa_select

    repo     = PgResidentRepository(db)
    resident = await repo.get_by_id(resident_id)
    if not resident:
        raise HTTPException(status_code=404, detail="Warga tidak ditemukan")

    # Verify Ketua RT is managing this resident's RT group
    from app.modules.iam.infrastructure.models import RTGroupModel
    rt_result = await db.execute(
        sa_select(RTGroupModel).where(
            RTGroupModel.admin_user_id == _uuid.UUID(current_user["user_id"])
        )
    )
    rt_group = rt_result.scalar_one_or_none()
    if not rt_group or rt_group.id != resident.rt_group_id:
        raise HTTPException(
            status_code=403,
            detail="Anda tidak memiliki akses ke data warga ini"
        )

    # Fetch changer's name for the log
    changer = await db.get(UserModel, _uuid.UUID(current_user["user_id"]))
    changer_name = changer.full_name if changer else "Unknown"

    # ── Build diff — only process fields that were explicitly provided ────
    # Use model_fields_set to detect which fields were in the request body.
    # This correctly distinguishes "not sent" from "sent as null".
    provided = body.model_fields_set
    now      = datetime.now(timezone.utc)
    logs     = []

    def get_old_value(field: str) -> str | None:
        """Get current string representation of a field for the log."""
        val = getattr(resident, field, None)
        if val is None:
            return None
        if hasattr(val, "value"):   # enum
            return val.value
        return str(val)

    def coerce_enum(field: str, raw: str | None):
        """Safely coerce a string to the correct domain enum."""
        if raw is None:
            return None
        enum_map = {
            "jenis_kelamin":       JenisKelamin,
            "agama":               Agama,
            "pekerjaan":           Pekerjaan,
            "status_kawin":        StatusKawin,
            "status_tinggal":      StatusTinggal,
            "status_keluarga":     StatusKeluarga,
            "pendidikan_terakhir": PendidikanTerakhir,
            "kewarganegaraan":     Kewarganegaraan,
            "hubungan_dengan_kk":  HubunganDenganKK,
        }
        if field in enum_map:
            try:
                return enum_map[field](raw)
            except ValueError:
                raise HTTPException(
                    status_code=422,
                    detail=f"Nilai '{raw}' tidak valid untuk field {field}"
                )
        return raw

    # Build kwargs for update_profile() + collect log entries
    update_kwargs: dict = {}

    for field in FIELD_LABELS:
        if field not in provided:
            continue   # not in request — skip

        raw_new   = getattr(body, field, None)
        old_value = get_old_value(field)

        # Coerce enum fields
        coerced = coerce_enum(field, raw_new)
        update_kwargs[field] = coerced

        # Determine new_value string for log
        if coerced is None:
            new_value = None
        elif hasattr(coerced, "value"):
            new_value = coerced.value
        elif isinstance(coerced, bool):
            new_value = "Ya" if coerced else "Tidak"
        else:
            new_value = str(coerced)

        # Only log if value actually changed
        old_str = "Ya" if old_value == "True" else ("Tidak" if old_value == "False" else old_value)
        if old_str != new_value:
            logs.append({
                "id":             _uuid.uuid4(),
                "resident_id":    resident_id,
                "rt_group_id":    resident.rt_group_id,
                "changed_by":     _uuid.UUID(current_user["user_id"]),
                "changed_by_role": current_user.get("role", "ketua_rt"),
                "changed_by_name": changer_name,
                "resident_name":  resident.full_name,
                "field_name":     field,
                "field_label":    FIELD_LABELS[field],
                "old_value":      old_str,
                "new_value":      new_value,
                "changed_at":     now,
            })

    if not update_kwargs:
        return {"message": "Tidak ada perubahan", "changed_fields": 0}

    # ── Apply update via domain entity ────────────────────────────────────
    resident.update_profile(**update_kwargs)
    await repo.save(resident)

    # ── Write audit log entries ───────────────────────────────────────────
    # Direct insert — not through repository pattern (pure infrastructure)
    if logs:
        from app.modules.warga.infrastructure.models import \
            ResidentChangeLogModel
        for log in logs:
            db.add(ResidentChangeLogModel(**log))

    await db.commit()

    return {
        "message":        f"Data {resident.full_name} berhasil diperbarui",
        "changed_fields": len(logs),
        "changes":        [
            {
                "field":     l["field_name"],
                "label":     l["field_label"],
                "old_value": l["old_value"],
                "new_value": l["new_value"],
            }
            for l in logs
        ],
    }


# ═══════════════════════════════════════════════════════════════════════════════
# GET /warga/{resident_id}/change-log
# Returns audit trail for a specific resident
# ═══════════════════════════════════════════════════════════════════════════════

@router.get("/warga/{resident_id}/change-log", response_model=list[ChangeLogEntry], tags=["Warga"])
async def get_resident_change_log(
    resident_id:  UUID,
    limit:        int  = 20,
    current_user: dict = Depends(require_admin),
    db:           AsyncSession = Depends(get_db),
):
    """
    Returns last N changes to a resident profile.
    Sorted newest-first. Used by the change log tab in KKDetailModal.
    """
    from app.modules.warga.infrastructure.models import ResidentChangeLogModel
    from sqlalchemy import desc as sa_desc
    from sqlalchemy import select as sa_select

    result = await db.execute(
        sa_select(ResidentChangeLogModel)
        .where(ResidentChangeLogModel.resident_id == resident_id)
        .order_by(sa_desc(ResidentChangeLogModel.changed_at))
        .limit(limit)
    )
    rows = result.scalars().all()

    return [
        {
            "id":              str(r.id),
            "field_name":      r.field_name,
            "field_label":     r.field_label,
            "old_value":       r.old_value,
            "new_value":       r.new_value,
            "changed_by":      str(r.changed_by),
            "changed_by_name": r.changed_by_name,
            "changed_by_role": r.changed_by_role,
            "resident_name":   r.resident_name,
            "changed_at":      r.changed_at.isoformat(),
        }
        for r in rows
    ]

def _stringify(value) -> str | None:
    """Convert any field value to string for storage in old_value/new_value."""
    if value is None:
        return None
    if hasattr(value, "isoformat"):  # date/datetime
        return value.isoformat()
    if hasattr(value, "value"):      # enum
        return value.value
    if isinstance(value, bool):
        return "Ya" if value else "Tidak"
    return str(value)


def _change_request_to_item(r, resident_name: str) -> ChangeRequestItem:
    return ChangeRequestItem(
        id                = r.id,
        resident_id       = r.resident_id,
        resident_name     = resident_name,
        requested_by      = r.requested_by,
        requested_by_name = r.requested_by_name,
        field_name        = r.field_name,
        field_label       = r.field_label,
        old_value         = r.old_value,
        new_value         = r.new_value,
        status            = r.status,
        reviewed_by_name  = r.reviewed_by_name,
        reviewed_at       = r.reviewed_at.isoformat() if r.reviewed_at else None,
        rejection_reason  = r.rejection_reason,
        created_at        = r.created_at.isoformat(),
    )


# ═══════════════════════════════════════════════════════════════════════════════
# POST /warga/me/change-requests
# Warga submits proposed field changes — pending Ketua RT approval
# ═══════════════════════════════════════════════════════════════════════════════

@router.post("/warga/me/change-requests", response_model=SubmitChangeRequestResponse,
              status_code=201, tags=["Warga"])
async def submit_change_request(
    body:         SubmitChangeRequestBody,
    current_user: dict = Depends(get_current_user),
    db:           AsyncSession = Depends(get_db),
):
    """
    Warga proposes changes to their own profile. One pending request is
    created per changed field. ResidentModel is NOT modified until a
    Ketua RT approves each request via /warga/change-requests/{id}/review.
    """
    user_id   = _uuid.UUID(current_user["user_id"])
    user_name = current_user.get("full_name", "")

    repo     = PgResidentRepository(db)
    resident = await repo.get_by_user_id(user_id)
    if not resident:
        raise HTTPException(status_code=404, detail="Data warga tidak ditemukan untuk akun ini")

    # Existing pending fields for this resident — block duplicate submissions
    existing_result = await db.execute(
        select(ResidentChangeRequestModel.field_name).where(
            ResidentChangeRequestModel.resident_id == resident.id,
            ResidentChangeRequestModel.status == "pending",
        )
    )
    pending_fields = {row[0] for row in existing_result.all()}

    payload = body.model_dump(exclude_unset=True)
    if not payload:
        raise HTTPException(status_code=422, detail="Tidak ada perubahan yang diajukan")

    created: list[ResidentChangeRequestModel] = []
    skipped: list[str] = []

    for field_name, new_value in payload.items():
        if field_name not in FIELD_LABELS:
            continue

        if field_name in pending_fields:
            skipped.append(field_name)
            continue

        old_str = _stringify(getattr(resident, field_name, None))
        new_str = _stringify(new_value)

        if old_str == new_str:
            continue  # no-op

        req = ResidentChangeRequestModel(
            id                = _uuid.uuid4(),
            resident_id       = resident.id,
            rt_group_id       = resident.rt_group_id,
            requested_by      = user_id,
            requested_by_name = user_name,
            field_name        = field_name,
            field_label       = FIELD_LABELS[field_name],
            old_value         = old_str,
            new_value         = new_str,
            status            = "pending",
        )
        db.add(req)
        created.append(req)

    if not created:
        msg = (
            "Semua field sudah memiliki permintaan yang menunggu persetujuan"
            if skipped else
            "Tidak ada perubahan terdeteksi dari data saat ini"
        )
        return SubmitChangeRequestResponse(created_count=0, requests=[], message=msg)

    await db.commit()
    for r in created:
        await db.refresh(r)

    msg = f"{len(created)} permintaan perubahan diajukan, menunggu persetujuan Ketua RT"
    if skipped:
        msg += f" ({len(skipped)} field dilewati karena sudah ada permintaan tertunda)"

    return SubmitChangeRequestResponse(
        created_count = len(created),
        requests      = [_change_request_to_item(r, resident.full_name) for r in created],
        message       = msg,
    )


# ═══════════════════════════════════════════════════════════════════════════════
# GET /warga/me/change-requests
# Warga views own request history
# ═══════════════════════════════════════════════════════════════════════════════

@router.get("/warga/me/change-requests", response_model=list[ChangeRequestItem], tags=["Warga"])
async def get_my_change_requests(
    current_user: dict = Depends(get_current_user),
    db:           AsyncSession = Depends(get_db),
):
    """Returns all change requests submitted by the logged-in warga, newest first."""
    user_id = _uuid.UUID(current_user["user_id"])

    repo     = PgResidentRepository(db)
    resident = await repo.get_by_user_id(user_id)
    if not resident:
        raise HTTPException(status_code=404, detail="Data warga tidak ditemukan untuk akun ini")

    result = await db.execute(
        select(ResidentChangeRequestModel)
        .where(ResidentChangeRequestModel.resident_id == resident.id)
        .order_by(desc(ResidentChangeRequestModel.created_at))
    )
    rows = result.scalars().all()
    return [_change_request_to_item(r, resident.full_name) for r in rows]


# ═══════════════════════════════════════════════════════════════════════════════
# GET /warga/change-requests/pending
# Ketua RT review queue — all pending self-edit requests for their RT
# ═══════════════════════════════════════════════════════════════════════════════

@router.get("/warga/change-requests/pending", response_model=list[ChangeRequestItem], tags=["Warga"])
async def get_pending_change_requests(
    current_user: dict = Depends(require_admin),
    db:           AsyncSession = Depends(get_db),
):
    """Returns all pending self-edit requests for the Ketua RT's RT group, oldest first."""
    from app.modules.iam.infrastructure.models import RTGroupModel

    user_id = _uuid.UUID(current_user["user_id"])

    rt_result = await db.execute(
        select(RTGroupModel).where(RTGroupModel.admin_user_id == user_id)
    )
    rt_group = rt_result.scalar_one_or_none()
    if not rt_group:
        raise HTTPException(status_code=404, detail="RT group tidak ditemukan")

    result = await db.execute(
        select(ResidentChangeRequestModel, ResidentModel.full_name)
        .join(ResidentModel, ResidentChangeRequestModel.resident_id == ResidentModel.id)
        .where(
            ResidentChangeRequestModel.rt_group_id == rt_group.id,
            ResidentChangeRequestModel.status == "pending",
        )
        .order_by(ResidentChangeRequestModel.created_at.asc())
    )
    rows = result.all()
    return [_change_request_to_item(r, resident_name) for r, resident_name in rows]


# ═══════════════════════════════════════════════════════════════════════════════
# PATCH /warga/change-requests/{request_id}/review
# Ketua RT approve or reject a single field-change request
# ═══════════════════════════════════════════════════════════════════════════════

@router.patch("/warga/change-requests/{request_id}/review", tags=["Warga"])
async def review_change_request(
    request_id:   UUID,
    body:         ReviewChangeRequestBody,
    current_user: dict = Depends(require_admin),
    db:           AsyncSession = Depends(get_db),
):
    """
    Approve or reject a warga's proposed field change.

    On approve:
      1. Apply new_value to ResidentModel
      2. If field is full_name or phone — ALSO apply to UserModel
         (keeps users table and residents table in sync, replicating
         the dual-write that PATCH /users/me/profile used to do)
      3. Write entry to resident_change_logs (changed_by_role="warga")
      4. Mark request approved

    On reject:
      1. Mark request rejected with reason
      2. Nothing is modified
    """
    from app.modules.iam.infrastructure.models import RTGroupModel, UserModel
    from app.modules.iam.infrastructure.repository import PgUserRepository

    if body.action not in ("approve", "reject"):
        raise HTTPException(status_code=422, detail="Action harus 'approve' atau 'reject'")
    if body.action == "reject" and not body.rejection_reason:
        raise HTTPException(status_code=422, detail="rejection_reason wajib diisi saat menolak")

    admin_id   = _uuid.UUID(current_user["user_id"])
    admin_name = current_user.get("full_name", "")

    rt_result = await db.execute(
        select(RTGroupModel).where(RTGroupModel.admin_user_id == admin_id)
    )
    rt_group = rt_result.scalar_one_or_none()
    if not rt_group:
        raise HTTPException(status_code=404, detail="RT group tidak ditemukan")

    req = await db.get(ResidentChangeRequestModel, request_id)
    if not req:
        raise HTTPException(status_code=404, detail="Permintaan tidak ditemukan")
    if req.rt_group_id != rt_group.id:
        raise HTTPException(status_code=403, detail="Permintaan ini bukan milik RT Anda")
    if req.status != "pending":
        raise HTTPException(status_code=409,
            detail=f"Permintaan sudah {req.status}, tidak bisa direview ulang")

    now = datetime.now(timezone.utc)

    if body.action == "reject":
        req.status           = "rejected"
        req.reviewed_by      = admin_id
        req.reviewed_by_name = admin_name
        req.reviewed_at      = now
        req.rejection_reason = body.rejection_reason
        await db.commit()
        return {"message": "Permintaan ditolak", "request_id": str(request_id)}

    # ── APPROVE ──────────────────────────────────────────────────────────
    resident = await db.get(ResidentModel, req.resident_id)
    if not resident:
        raise HTTPException(status_code=404, detail="Data warga tidak ditemukan")

    # Coerce stored string back to correct type for date fields
    new_value_raw: object = req.new_value
    if req.field_name == "tanggal_lahir" and req.new_value:
        from datetime import date as _date
        new_value_raw = _date.fromisoformat(req.new_value)

    # Coerce enum fields back via domain entities
    if req.field_name in (
        "jenis_kelamin", "agama", "pekerjaan", "status_kawin", "status_tinggal",
        "status_keluarga", "pendidikan_terakhir", "kewarganegaraan", "hubungan_dengan_kk",
    ) and req.new_value is not None:
        from app.modules.warga.domain.entities import (Agama, HubunganDenganKK,
                                                       JenisKelamin,
                                                       Kewarganegaraan,
                                                       Pekerjaan,
                                                       PendidikanTerakhir,
                                                       StatusKawin,
                                                       StatusKeluarga,
                                                       StatusTinggal)
        enum_map = {
            "jenis_kelamin":       JenisKelamin,
            "agama":               Agama,
            "pekerjaan":           Pekerjaan,
            "status_kawin":        StatusKawin,
            "status_tinggal":      StatusTinggal,
            "status_keluarga":     StatusKeluarga,
            "pendidikan_terakhir": PendidikanTerakhir,
            "kewarganegaraan":     Kewarganegaraan,
            "hubungan_dengan_kk":  HubunganDenganKK,
        }
        try:
            new_value_raw = enum_map[req.field_name](req.new_value)
        except ValueError:
            raise HTTPException(status_code=422,
                detail=f"Nilai '{req.new_value}' tidak valid untuk field {req.field_name}")

    # ── full_name / phone — special case: also sync UserModel ────────────
    if req.field_name == "phone":
        # Uniqueness check — mirrors old PATCH /users/me/profile behaviour
        user_repo = PgUserRepository(db)
        existing  = await user_repo.get_by_phone(req.new_value)
        if existing and existing.id != req.requested_by:
            raise HTTPException(
                status_code=409,
                detail="Nomor HP sudah digunakan akun lain — permintaan tidak dapat disetujui"
            )

    if req.field_name in ("full_name", "phone"):
        user = await db.get(UserModel, req.requested_by)
        if user:
            setattr(user, req.field_name, new_value_raw)

    # Capture current value as old_value (resident may have changed since request)
    log_old_value = _stringify(getattr(resident, req.field_name, None))

    # Apply via domain entity — keeps update_profile() validation in one place
    resident.update_profile(**{req.field_name: new_value_raw})

    repo = PgResidentRepository(db)
    await repo.save(resident)

    # Audit log — attributed to the warga who requested it
    db.add(ResidentChangeLogModel(
        id              = _uuid.uuid4(),
        resident_id     = resident.id,
        rt_group_id     = req.rt_group_id,
        changed_by      = req.requested_by,
        changed_by_role = "warga",
        changed_by_name = req.requested_by_name,
        resident_name   = resident.full_name,
        field_name      = req.field_name,
        field_label     = req.field_label,
        old_value       = log_old_value,
        new_value       = req.new_value,
    ))

    req.status           = "approved"
    req.reviewed_by      = admin_id
    req.reviewed_by_name = admin_name
    req.reviewed_at      = now

    await db.commit()

    return {
        "message":    f"Perubahan {req.field_label} disetujui dan diterapkan",
        "request_id": str(request_id),
    }


@router.get("/warga/import/template", tags=["Warga"])
async def download_import_template(
    current_user: dict = Depends(require_admin),
):
    """
    GET /warga/import/template
    Returns a pre-filled .xlsx template Ketua RT can download and fill in.
    Row 1 = headers, Row 2 = example row (greyed out in Excel).
    """
    import io

    import openpyxl
    from fastapi.responses import StreamingResponse
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data Warga"

    headers = [
        ("nama_lengkap",       "Nama Lengkap *",        20),
        ("no_whatsapp",        "No. WhatsApp *",         18),
        ("nik",                "NIK",                    18),
        ("no_kk",              "No. KK",                 18),
        ("tanggal_lahir",      "Tanggal Lahir (YYYY-MM-DD)", 22),
        ("tempat_lahir",       "Tempat Lahir",           18),
        ("jenis_kelamin",      "Jenis Kelamin",          16),
        ("agama",              "Agama",                  12),
        ("pekerjaan",          "Pekerjaan",              22),
        ("status_kawin",       "Status Kawin",           16),
        ("status_tinggal",     "Status Tinggal",         16),
        ("status_keluarga",    "Status Keluarga",        16),
        ("alamat_ktp",         "Alamat KTP",             30),
        ("alamat_domisili",    "Alamat Domisili",        30),
        ("pendidikan_terakhir","Pendidikan Terakhir",    20),
        ("kewarganegaraan",    "Kewarganegaraan",        16),
        ("hubungan_dengan_kk", "Hubungan dengan KK",    20),
    ]

    # ── Header row ────────────────────────────────────────────────────────────
    header_fill     = PatternFill("solid", fgColor="1E3A5F")
    header_font     = Font(bold=True, color="FFFFFF", size=11)
    required_fill   = PatternFill("solid", fgColor="2563EB")

    for col_idx, (key, label, width) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=key)
        cell.font      = header_font
        cell.fill      = required_fill if key in ("nama_lengkap", "no_whatsapp") else header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 36

    # ── Example row ───────────────────────────────────────────────────────────
    example_data = [
        "Budi Santoso", "081234567890", "3171234567890001", "3171234567890001",
        "1990-05-21", "Jakarta", "LAKI-LAKI", "ISLAM",
        "KARYAWAN SWASTA", "KAWIN", "TETAP", "SUAMI",
        "Jl. Mawar No. 5 RT 001", "Jl. Mawar No. 5 RT 001",
        "S1", "WNI", "KEPALA KELUARGA",
    ]
    example_fill = PatternFill("solid", fgColor="F0F4FF")
    example_font = Font(italic=True, color="6B7280", size=10)

    for col_idx, value in enumerate(example_data, start=1):
        cell = ws.cell(row=2, column=col_idx, value=value)
        cell.fill      = example_fill
        cell.font      = example_font
        cell.alignment = Alignment(vertical="center")

    ws.row_dimensions[2].height = 20
    ws.freeze_panes = "A2"  # freeze header row

    # ── Save to bytes ─────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=template_import_warga.xlsx"},
    )


@router.post("/warga/import/preview", tags=["Warga"])
async def preview_import(
    current_user: dict = Depends(require_admin),
    file: UploadFile = File(..., description=".xlsx file — gunakan template RTMudah"),
):
    """
    POST /warga/import/preview
    Phase 1 — Upload .xlsx, validate every row, return preview.
    Nothing is written to the database.
    """
    from app.modules.warga.application.use_cases.import_residents import \
        parse_excel

    # ── File type guard ───────────────────────────────────────────────────────
    filename = file.filename or ""
    if not filename.lower().endswith(".xlsx"):
        raise HTTPException(
            status_code=422,
            detail="Hanya file .xlsx yang didukung. Gunakan template yang disediakan."
        )

    file_bytes = await file.read()

    if len(file_bytes) > 5 * 1024 * 1024:   # 5 MB hard limit
        raise HTTPException(
            status_code=422,
            detail="Ukuran file maksimal 5 MB"
        )

    try:
        result = parse_excel(file_bytes)
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))

    return result


@router.post("/warga/import/confirm", status_code=201, tags=["Warga"])
async def confirm_import(
    body:         dict,
    current_user: dict = Depends(require_admin),
    db:           AsyncSession = Depends(get_db),
):
    """
    POST /warga/import/confirm
    Phase 2 — Bulk-insert the validated rows returned by /preview.
    Body: { "rows": [...valid rows from preview response...] }
    """
    from app.modules.iam.infrastructure.models import RTGroupModel
    from app.modules.warga.application.use_cases.import_residents import \
        bulk_create
    from sqlalchemy import select as sa_select

    rows = body.get("rows", [])
    if not rows:
        raise HTTPException(status_code=422, detail="Tidak ada data untuk diimport")

    if len(rows) > 500:
        raise HTTPException(
            status_code=422,
            detail="Maksimal 500 baris per import. Pecah file menjadi beberapa bagian."
        )

    admin_id = _uuid.UUID(current_user["user_id"])

    rt_result = await db.execute(
        sa_select(RTGroupModel).where(RTGroupModel.admin_user_id == admin_id)
    )
    rt_group = rt_result.scalar_one_or_none()
    if not rt_group:
        raise HTTPException(status_code=404, detail="RT group tidak ditemukan")

    result = await bulk_create(
        valid_rows=rows,
        rt_group_id=rt_group.id,
        added_by_user_id=admin_id,
        db=db,
    )

    return {
        "message":     f"{result['imported']} warga berhasil diimport",
        "imported":    result["imported"],
        "failed":      result["failed"],
        "failed_rows": result["failed_rows"],
    }
